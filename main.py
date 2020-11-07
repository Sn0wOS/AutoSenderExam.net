
from PyQt5.QtWidgets import QApplication,\
    QFileDialog, QPushButton, QFileDialog, \
    QLabel, QTableView, QHBoxLayout, QVBoxLayout, QWidget,\
    QTextEdit, QProgressBar

from PyQt5.QtCore import Qt, QModelIndex
from PyQt5.QtCore import QAbstractTableModel
import sys
import os
import csv
import smtplib
import typing
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from time import sleep
from openpyxl import Workbook


class TableModel(QAbstractTableModel):
    EditableLocation = [  # remeber the col coynt from zero
        4,  # ExamFile col
        5
    ]

    def __init__(self, data):
        super(TableModel, self).__init__()
        self._data = data

    def data(self, index, role):
        if role == Qt.DisplayRole:
            # See below for the nested-list data structure.
            # .row() indexes into the outer list,
            # .column() indexes into the sub-list
            return self._data[index.row()][index.column()]

    def rowCount(self, index):
        # The length of the outer list.
        return len(self._data)

    def columnCount(self, index):
        # The following takes the first sub-list, and returns
        # the length (only works if all rows are an equal length)
        return len(self._data[0])

    def setData(self, index: QModelIndex, value: typing.Any, role: int) -> bool:
        # print("row:", index.row(), "col:", index.column(),
        #      'value:', type(value), value)
        try:
            if 0 <= int(value) <= 200:
                self._data[index.row()][index.column()] = value
            else:
                return True
        except ValueError as e:
            self._data[index.row()][index.column()] = 0
            pass

        return True

    def flags(self, index: QModelIndex) -> Qt.ItemFlags:
        retFalgDefult = Qt.ItemIsEnabled | Qt.ItemNeverHasChildren
        if index.column() in self.EditableLocation:
            return retFalgDefult | Qt.ItemIsEditable
        else:
            return retFalgDefult


class App(QWidget):
    NOT_SELECTED_STRING = "לא נבחר"
    NOTE_PLACE_HOLDER = "הכנס כאן הארה/הערה כללית עבור כלל התלמידים (רשות)"
    FROM_LABEL = "שם המורה !!!!!באנגלית בלבד!!!!!"
    EXAM_NAME_LABEL = "שם הבחינה"

    def __init__(self) -> None:
        super().__init__()
        self.title = "מערכת שליחת מיילים של קבצי בדיקה"
        self.left = 50
        self.top = 50
        self.width = 800
        self.height = 800
        self.csv_directory = ""
        self.answers_directory = ""
        self.dataTableSrc = []
        self.studentList = []
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        self.csv_load_btn = QPushButton("הוסף רשימת תלמדים")
        self.folder_load_btn = QPushButton("הוסף תיקיית קבצים")

        self.csv_file_name = QLabel()
        self.folder_dir_name = QLabel()

        self.credit = QLabel(
            "made by Shlomo Neuberer\n shlomo.dev@neuberger.co.il \n tel:+972-503305021")
        self.mailInfoTitle = QLabel("אזור עריכת תוכן המייל")
        self.noteLabel = QTextEdit()
        self.fromLabel = QTextEdit()
        self.ExamNamelable = QTextEdit()
        self.sendingProgbar = QProgressBar()
        self.sendingProgbar.show()
        self.sendingProgbar.reset()
        self.sendingProgbar.setMinimum(0)
        self.sendingProgbar.setTextVisible(True)
        self.sendingProgbar.hide()

        self.noteLabel.setPlaceholderText(self.NOTE_PLACE_HOLDER)
        self.fromLabel.setPlaceholderText(self.FROM_LABEL)
        self.ExamNamelable.setPlaceholderText(self.EXAM_NAME_LABEL)

        self.send_btn = QPushButton("שלח", self)

        self.table = QTableView(self)

        self.model = TableModel([])

        self.mainVerticalArea = QVBoxLayout()
        self.VBox_topArea = QVBoxLayout()
        self.HBOX_topArea = QHBoxLayout()
        self.VBox_mailInfo = QVBoxLayout()
        self.HBox_addStudent = QHBoxLayout()
        self.HBox_addFolderAnswer = QHBoxLayout()

        self.VBox_mailInfo.addWidget(self.mailInfoTitle)
        self.VBox_mailInfo.addWidget(self.noteLabel)
        self.VBox_mailInfo.addWidget(self.fromLabel)
        self.VBox_mailInfo.addWidget(self.ExamNamelable)
        self.VBox_mailInfo.addStretch(1)
        self.VBox_mailInfo.addWidget(self.credit)

        self.HBox_addStudent.addWidget(self.csv_load_btn)
        self.HBox_addStudent.addWidget(self.csv_file_name)

        self.HBox_addFolderAnswer.addWidget(self.folder_load_btn)
        self.HBox_addFolderAnswer.addWidget(self.folder_dir_name)

        self.VBox_topArea.addLayout(self.HBox_addStudent)
        self.VBox_topArea.addLayout(self.HBox_addFolderAnswer)

        self.HBOX_topArea.addLayout(self.VBox_topArea)
        self.HBOX_topArea.addSpacing(100)
        self.HBOX_topArea.addLayout(self.VBox_mailInfo)

        self.mainVerticalArea.addLayout(self.HBOX_topArea)

        self.sen_btn_progreesbar = QHBoxLayout()
        self.sen_btn_progreesbar.addWidget(self.send_btn)
        self.sen_btn_progreesbar.addWidget(self.sendingProgbar)
        self.mainVerticalArea.addLayout(self.sen_btn_progreesbar)
        self.mainVerticalArea.addWidget(self.table)

        self.csv_load_btn.setToolTip("רשימת תלמידים שהופק על ידי exam.net")
        self.folder_load_btn.setToolTip(
            "תיקייה המכילה את הקבצים של התשובות")
        self.folder_dir_name.setText(self.NOT_SELECTED_STRING)
        self.csv_file_name.setText(self.NOT_SELECTED_STRING)

        self.csv_load_btn.clicked.connect(self.open_dialog_get_students_csv)
        self.folder_load_btn.clicked.connect(self.open_dialog_get_answer_pdfs)

        self.send_btn.clicked.connect(
            lambda: self.send_mails(self.fromLabel.toPlainText(),
                                    self.ExamNamelable.toPlainText(),
                                    notes=self.noteLabel.toPlainText()
                                    )
        )

        self.setLayout(self.mainVerticalArea)

        self.show()

    def open_dialog_get_students_csv(self):
        fileDialog = QFileDialog()
        # getOpen file return tuple (fileName,filter)
        fileName, _ = fileDialog. getOpenFileName(
            self, "בחר קובץ רשימת תלמדים", "", "Files (*.csv)")
        if os.path.exists(fileName):
            self.csv_directory = fileName
            self.csv_file_name.setText(os.path.basename(fileName))
            self.updateList()
            self.table.setModel(TableModel(self.dataTableSrc))

    def open_dialog_get_answer_pdfs(self):
        fileDialog = QFileDialog()
        directory = fileDialog.getExistingDirectory(
            self, "הוסף תיקיית תשובות", "Desktop")
        if os.path.exists(directory):
            self.answers_directory = directory
            self.folder_dir_name.setText(os.path.basename(directory))
            self.updateList()
            self.table.setModel(TableModel(self.dataTableSrc))

    def __getStudentListFromCSV(self, csv_directory: str) -> list:
        csvList = []
        childDict = {"Id": "ID", 'FirstName': 'FirstName', 'LastName': 'LastName',
                     'Email': 'Email', 'file': 'Exam File', 'Grade': 'Grade'}

        try:
            with open(csv_directory, encoding="utf-8") as csvFile:
                reader = csv.reader(csvFile, delimiter=';')
                for row in reader:

                    splitted = row[0].split(" ")
                    childDict['Id'] = splitted[0]
                    childDict['FirstName'] = splitted[1]
                    childDict['LastName'] = row[1]
                    childDict['Email'] = row[2]
                    csvList.append(childDict.copy())
                    childDict['Exam File'] = ''
                    childDict['Grade'] = '54'

        except Exception as e:
            print(e)
            return None

        return csvList

    def __findFilesInDirectory(self, answers_directory, studentList) -> list:
        if answers_directory == "" or answers_directory == self.NOT_SELECTED_STRING:
            return studentList
        for _, _, fileNames in os.walk(answers_directory):
            for name in fileNames:  # itrate on all files in given directory
                for student in studentList:  # iterate on all student name
                    # test that sure and last name apper in the file
                    if student['FirstName'] in name and student['LastName'] in name:
                        # print('email:', student['Email'], 'file', name)
                        if student['file'] == "Exam File":
                            student['file'] = os.path.join(
                                answers_directory, name)
                        else:
                            student['file'] = student['file']+'&'+os.path.join(
                                answers_directory, name)
        return studentList

    def updateList(self):
        self.studentList = self.__getStudentListFromCSV(self.csv_directory)
        if isinstance(self.studentList, list):
            self.studentList = self.__findFilesInDirectory(
                self.answers_directory, self.studentList)
            if isinstance(self.studentList, list):
                self.dataTableSrc = self.list_of_dict_to_list(self.studentList)

    def send_mails(self, From, examName, notes=""):
        self.sendingProgbar.setMaximum(len(self.studentList))
        self.sendingProgbar.show()
        countRow = 0
        for student in self.studentList:
            grade = self.dataTableSrc[countRow][5]
            fullname = student['FirstName']+" "+student["LastName"]
            countRow = countRow + 1
            bodyBaseString = "{}\nשלום ל {} \nלהלן מצורפת הבחינה שלך ב{}\nהציון שלך {}\nבהצלחה להבא מ\n{}".format(
                notes, fullname, examName, grade, From)
            # print(bodyBaseString)
            self.__send_mail(
                From, student['Email'], examName, bodyBaseString, filename=student['file'])
            self.sendingProgbar.setValue(countRow)
        # self.sendingProgbar.hide()
        self.__exportToFile()
        sleep(0.25)
        self.sendingProgbar.hide()

    def __send_mail(self, sent_from, to, subject, body, filename=None):
        gmail_user = 'userName'
        gmail_password = 'PassWord'
        try:
            msg = MIMEMultipart()
            msg['From'] = sent_from
            msg['To'] = to
            msg['Date'] = formatdate(localtime=True)
            msg['Subject'] = subject

            msg.attach(MIMEText(body))
            if type(filename) is str:
                if '&' in filename:
                    listFiles = filename.split('&')
                    for name in listFiles:
                        if name != "":
                            with open(name, "rb") as f:
                                part = MIMEApplication(
                                    f.read(), _subtype="pdf")
                            part.add_header('Content-Disposition', 'attachment',
                                            filename=os.path.basename(name))
                            msg.attach(part)
                else:
                    with open(filename, "rb") as f:
                        part = MIMEApplication(f.read(), _subtype="pdf")
                    part.add_header('Content-Disposition', 'attachment',
                                    filename=os.path.basename(filename))
                    msg.attach(part)
            else:
                return
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
            server.ehlo()
            server.login(gmail_user, gmail_password)
            server.sendmail(sent_from, to, msg.as_string())
            server.close()

            print('Email sent!')
        except Exception as e:
            print('Something went wrong...\n', e)

    def list_of_dict_to_list(self, src: list) -> list:
        ret = []
        for s in src:
            temp = []
            for _, value in s.items():
                temp.append(value)

            ret.append(temp)
        return ret

    def __exportToFile(self):
        work = Workbook()
        sheet = work.create_sheet("grades Sheet")
        rowCount = 1
        for student in self.studentList:
            sheet.cell(row=rowCount, column=1, value=student['Id'])
            sheet.cell(row=rowCount, column=2, value=student['FirstName'])
            sheet.cell(row=rowCount, column=3, value=student['LastName'])
            sheet.cell(row=rowCount, column=4, value=student['Grade'])
            rowCount = rowCount+1
        work.save(os.path.join(self.answers_directory, "export.xls"))


if __name__ == "__main__":
    app = QApplication([])
    ex = App()
    sys.exit(app.exec_())
